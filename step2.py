import os
import re
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime


def split_into_sections(pairs):
    sections, curr = [], {}
    for q, val in pairs:
        if q == 1 and curr:
            sections.append(curr)
            curr = {}
        curr[q] = val
    if curr:
        sections.append(curr)
    return sections


def clean_latex(s: str) -> str:
    fixes = [
        (r'\\frac\{([^}]+)\}\{([^}]+)\}', r'\1/\2'),
        (r'\\ldots', '...'),
        (r'\\times', '×'),
        (r'\\rightarrow', '→'),
        (r'\\le\b', '≤'),
        (r'\\ge\b', '≥'),
        (r'\\quad', ' '),
        (r'\$(.*?)\$', r'\1'),
        (r'\\', '')
    ]
    for pat, rep in fixes:
        s = re.sub(pat, rep, s)
    return s.strip()


def process_step2(ans_md_path: str, sol_md_path: str, input_xlsx: str, output_path: str = None) -> str:
    """
    Reads answer-key .md, solution .md, merges into input_xlsx, writes to new Excel.
    """
    # Parse answer key
    ans_txt = Path(ans_md_path).read_text(encoding='utf-8')
    ans_pairs = []
    for line in ans_txt.splitlines():
        m = re.match(r'^\s*(\d+)\.\s*\(\s*([abcd])\s*\)', line, flags=re.IGNORECASE)
        if m:
            ans_pairs.append((int(m.group(1)), m.group(2).lower()))
        else:
            m2 = re.match(r'^\s*(\d+)\.\s*(.+)$', line)
            if m2 and not re.fullmatch(r'[abcd]', m2.group(2), flags=re.IGNORECASE):
                ans_pairs.append((int(m2.group(1)), m2.group(2).strip()))
    answer_sections = split_into_sections(ans_pairs)

    # Parse solutions
    sol_txt = Path(sol_md_path).read_text(encoding='utf-8')
    sol_pairs, cur_q, cur_lines = [], None, []
    for line in sol_txt.splitlines(keepends=True):
        m = re.match(r'^\s*(\d+)\.\s*(.*)', line)
        if m:
            if cur_q is not None:
                sol_pairs.append((cur_q, ''.join(cur_lines).strip()))
            cur_q = int(m.group(1))
            cur_lines = [m.group(2) + '\n']
        elif cur_q is not None:
            cur_lines.append(line)
    if cur_q is not None:
        sol_pairs.append((cur_q, ''.join(cur_lines).strip()))
    solution_sections = split_into_sections(sol_pairs)

    # Load and merge
    wb = load_workbook(input_xlsx)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    q_col = hdr.index('Question No')+1
    ans_col = hdr.index('Answer')+1
    sol_col = hdr.index('Explanation')+1

    prev_q, sec_idx = None, 0
    for row in range(2, ws.max_row+1):
        val = ws.cell(row, q_col).value
        try:
            qn = int(val)
        except:
            continue
        if prev_q is not None and qn == 1:
            sec_idx += 1
        prev_q = qn
        if sec_idx >= len(answer_sections):
            break
        a = answer_sections[sec_idx].get(qn, '')
        e = solution_sections[sec_idx].get(qn, '')
        ws.cell(row, ans_col, value=a)
        ws.cell(row, sol_col, value=clean_latex(e))

    if not output_path:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        base = os.path.dirname(input_xlsx)
        output_path = os.path.join(base, f"2_{ts}.xlsx")
    wb.save(output_path)
    return output_path